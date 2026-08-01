import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def afficher_en_tete():
    """Affiche l'en-tête du programme"""
    print("="*70)
    print("   CALCUL DES MOYENNES AVEC COEFFICIENTS - CLASSE MP2I")
    print("="*70)
    print()

def charger_fichier_excel(chemin_fichier):
    """
    Charge le fichier Excel et retourne le DataFrame
    
    Args:
        chemin_fichier (str): Chemin vers le fichier Excel
    
    Returns:
        pandas.DataFrame: Données chargées
    """
    try:
        df = pd.read_excel(chemin_fichier)
        print(f"✓ Fichier chargé: {chemin_fichier}")
        print(f"✓ Nombre de lignes: {len(df)}")
        print(f"✓ Nombre de colonnes: {len(df.columns)}")
        return df
    except Exception as e:
        print(f"✗ Erreur lors du chargement: {e}")
        return None

def afficher_structure(df):
    """
    Affiche la structure du fichier (colonnes et premières lignes)
    
    Args:
        df (pandas.DataFrame): DataFrame à analyser
    """
    print("\n" + "="*70)
    print("STRUCTURE DU FICHIER:")
    print("="*70)
    print("\nColonnes disponibles:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i}. {col}")
    
    print("\n" + "-"*70)
    print("Aperçu des 3 premières lignes:")
    print("-"*70)
    print(df.head(3).to_string())
    print()

def identifier_colonnes_notes(df):
    """
    Identifie automatiquement les colonnes contenant des notes
    
    Args:
        df (pandas.DataFrame): DataFrame à analyser
    
    Returns:
        list: Liste des noms de colonnes de notes
    """
    colonnes_notes = []
    colonnes_info = ['Numéro', 'Nom', 'Prénoms', 'Prénom', 'Classe', 'Moyenne']
    
    for col in df.columns:
        col_lower = str(col).lower()
        # Exclure les colonnes d'information
        est_info = any(info.lower() in col_lower for info in colonnes_info)
        
        if not est_info:
            # Vérifier si la colonne contient des données numériques
            try:
                valeurs_numeriques = pd.to_numeric(df[col], errors='coerce')
                if valeurs_numeriques.notna().sum() > 0:
                    colonnes_notes.append(col)
            except:
                pass
    
    return colonnes_notes

def demander_coefficients(colonnes_notes):
    """
    Demande à l'utilisateur de saisir les coefficients pour chaque colonne
    
    Args:
        colonnes_notes (list): Liste des colonnes de notes
    
    Returns:
        dict: Dictionnaire {nom_colonne: coefficient}
    """
    print("\n" + "="*70)
    print("SAISIE DES COEFFICIENTS")
    print("="*70)
    print("\nVeuillez entrer le coefficient pour chaque évaluation:")
    print("(Appuyez sur Entrée pour utiliser le coefficient par défaut)")
    print()
    
    coefficients = {}
    
    for col in colonnes_notes:
        while True:
            try:
                # Déterminer le coefficient par défaut selon le type
                col_lower = str(col).lower()
                if 'ds' in col_lower or 'devoir' in col_lower:
                    coef_defaut = 2
                elif 'interro' in col_lower or 'interrogation' in col_lower:
                    coef_defaut = 1
                elif 'tp' in col_lower or 'pratique' in col_lower:
                    coef_defaut = 1
                elif 'partici' in col_lower:
                    coef_defaut = 0.5
                else:
                    coef_defaut = 1
                
                reponse = input(f"  {col} (défaut: {coef_defaut}): ").strip()
                
                if reponse == "":
                    coefficients[col] = coef_defaut
                    break
                else:
                    coef = float(reponse.replace(',', '.'))
                    if coef >= 0:
                        coefficients[col] = coef
                        break
                    else:
                        print("    ⚠ Le coefficient doit être positif. Réessayez.")
            except ValueError:
                print("    ⚠ Valeur invalide. Entrez un nombre. Réessayez.")
    
    return coefficients

def afficher_recapitulatif_coefficients(coefficients):
    """
    Affiche un récapitulatif des coefficients saisis
    
    Args:
        coefficients (dict): Dictionnaire des coefficients
    """
    print("\n" + "="*70)
    print("RÉCAPITULATIF DES COEFFICIENTS")
    print("="*70)
    somme_coef = sum(coefficients.values())
    
    for col, coef in coefficients.items():
        pourcentage = (coef / somme_coef * 100) if somme_coef > 0 else 0
        print(f"  {col:<40} Coef: {coef:>5.1f}  ({pourcentage:>5.1f}%)")
    
    print("-"*70)
    print(f"  {'TOTAL':<40} Coef: {somme_coef:>5.1f}  (100.0%)")
    print()

def calculer_moyenne_ponderee(notes, coefficients):
    """
    Calcule la moyenne pondérée pour une série de notes
    
    Args:
        notes (dict): Dictionnaire {colonne: note}
        coefficients (dict): Dictionnaire {colonne: coefficient}
    
    Returns:
        float: Moyenne pondérée
    """
    somme_ponderee = 0
    somme_coefficients = 0
    
    for col, note in notes.items():
        if pd.notna(note) and col in coefficients:
            somme_ponderee += note * coefficients[col]
            somme_coefficients += coefficients[col]
    
    if somme_coefficients > 0:
        return somme_ponderee / somme_coefficients
    else:
        return None

def calculer_moyennes_toute_classe(df, colonnes_notes, coefficients):
    """
    Calcule les moyennes pour tous les étudiants
    
    Args:
        df (pandas.DataFrame): DataFrame avec les notes
        colonnes_notes (list): Liste des colonnes de notes
        coefficients (dict): Dictionnaire des coefficients
    
    Returns:
        pandas.Series: Série contenant les moyennes calculées
    """
    moyennes = []
    
    for index, row in df.iterrows():
        notes = {col: row[col] for col in colonnes_notes}
        moyenne = calculer_moyenne_ponderee(notes, coefficients)
        moyennes.append(moyenne)
    
    return pd.Series(moyennes)

def creer_fichier_excel_avec_moyennes(df, colonnes_notes, coefficients, fichier_sortie):
    """
    Crée un fichier Excel avec les moyennes calculées et formules
    
    Args:
        df (pandas.DataFrame): DataFrame original
        colonnes_notes (list): Liste des colonnes de notes
        coefficients (dict): Dictionnaire des coefficients
        fichier_sortie (str): Nom du fichier de sortie
    """
    # Créer une copie du DataFrame
    df_resultat = df.copy()
    
    # Calculer les moyennes
    df_resultat['Moyenne Calculée'] = calculer_moyennes_toute_classe(df, colonnes_notes, coefficients)
    
    # Sauvegarder temporairement en Excel
    temp_file = 'temp_moyennes.xlsx'
    df_resultat.to_excel(temp_file, index=False)
    
    # Charger avec openpyxl pour ajouter le formatage et les formules
    wb = load_workbook(temp_file)
    sheet = wb.active
    
    # Ajouter une ligne avec les coefficients
    sheet.insert_rows(2)
    sheet.cell(row=2, column=1, value='Coefficients')
    sheet.cell(row=2, column=1).font = Font(bold=True, italic=True, color='0000FF')
    
    # Trouver les colonnes de notes et ajouter les coefficients
    col_indices = {}
    for i, col_name in enumerate(df.columns, 1):
        if col_name in colonnes_notes:
            col_indices[col_name] = i
            sheet.cell(row=2, column=i, value=coefficients[col_name])
            cell = sheet.cell(row=2, column=i)
            cell.font = Font(bold=True, color='0000FF')
            cell.alignment = Alignment(horizontal='center')
    
    # Formater l'en-tête
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', start_color='4472C4')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajouter une colonne avec formule Excel pour la moyenne
    col_moyenne_formule = len(df.columns) + 2
    col_letter_moyenne = chr(64 + col_moyenne_formule) if col_moyenne_formule <= 26 else 'A' + chr(64 + col_moyenne_formule - 26)
    
    sheet.cell(row=1, column=col_moyenne_formule, value='Moyenne (Formule)')
    cell = sheet.cell(row=1, column=col_moyenne_formule)
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', start_color='2E7D32')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Créer les formules SUMPRODUCT pour chaque étudiant
    for ligne in range(3, sheet.max_row + 1):
        # Construire la liste des cellules de notes et coefficients
        cellules_notes = []
        cellules_coefs = []
        
        for col_name, col_idx in col_indices.items():
            col_letter = chr(64 + col_idx) if col_idx <= 26 else 'A' + chr(64 + col_idx - 26)
            cellules_notes.append(f"{col_letter}{ligne}")
            cellules_coefs.append(f"{col_letter}2")
        
        # Créer la formule
        if cellules_notes:
            notes_str = ','.join(cellules_notes)
            coefs_str = ','.join(cellules_coefs)
            formule = f'=SUMPRODUCT({{{notes_str}}},{{{coefs_str}}})/SUM({{{coefs_str}}})'
            
            cell = sheet.cell(row=ligne, column=col_moyenne_formule, value=formule)
            cell.fill = PatternFill('solid', start_color='E8F5E9')
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '0.00'
    
    # Formater la colonne moyenne calculée (Python)
    col_moyenne_calc = len(df.columns) + 1
    for ligne in range(3, sheet.max_row + 1):
        cell = sheet.cell(row=ligne, column=col_moyenne_calc)
        cell.fill = PatternFill('solid', start_color='FFF9C4')
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.00'
    
    # Ajuster les largeurs des colonnes
    for column in sheet.columns:
        max_length = 0
        col_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 35)
        sheet.column_dimensions[col_letter].width = adjusted_width
    
    # Figer les volets
    sheet.freeze_panes = 'D3'
    
    # Ajouter des bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Sauvegarder
    wb.save(fichier_sortie)
    print(f"\n✓ Fichier Excel créé: {fichier_sortie}")

def afficher_statistiques(df, colonnes_notes):
    """
    Affiche les statistiques sur les notes
    
    Args:
        df (pandas.DataFrame): DataFrame avec les moyennes
        colonnes_notes (list): Liste des colonnes de notes
    """
    print("\n" + "="*70)
    print("STATISTIQUES DE LA CLASSE")
    print("="*70)
    
    if 'Moyenne Calculée' in df.columns:
        moyennes = df['Moyenne Calculée'].dropna()
        
        if len(moyennes) > 0:
            print(f"\n  Nombre d'étudiants: {len(moyennes)}")
            print(f"  Moyenne générale de la classe: {moyennes.mean():.2f}/20")
            print(f"  Note la plus haute: {moyennes.max():.2f}/20")
            print(f"  Note la plus basse: {moyennes.min():.2f}/20")
            print(f"  Écart-type: {moyennes.std():.2f}")
            
            # Distribution
            print("\n  Distribution des moyennes:")
            print(f"    >= 16: {(moyennes >= 16).sum()} étudiants")
            print(f"    14-16: {((moyennes >= 14) & (moyennes < 16)).sum()} étudiants")
            print(f"    12-14: {((moyennes >= 12) & (moyennes < 14)).sum()} étudiants")
            print(f"    10-12: {((moyennes >= 10) & (moyennes < 12)).sum()} étudiants")
            print(f"    < 10:  {(moyennes < 10).sum()} étudiants")

def programme_principal():
    """Fonction principale qui orchestre l'exécution du programme"""
    
    afficher_en_tete()
    
    # Chemin du fichier (à modifier selon votre fichier)
    fichier_entree = 'LISTE_DE_CLASSE_MP2I.xlsx'
    fichier_sortie = 'LISTE_MP2I_AVEC_MOYENNES.xlsx'
    
    # 1. Charger le fichier
    df = charger_fichier_excel(fichier_entree)
    if df is None:
        return
    
    # 2. Afficher la structure
    afficher_structure(df)
    
    # 3. Identifier les colonnes de notes
    colonnes_notes = identifier_colonnes_notes(df)
    
    if not colonnes_notes:
        print("✗ Aucune colonne de notes détectée!")
        return
    
    print(f"\n✓ {len(colonnes_notes)} colonnes de notes détectées")
    
    # 4. Demander les coefficients
    coefficients = demander_coefficients(colonnes_notes)
    
    # 5. Afficher le récapitulatif
    afficher_recapitulatif_coefficients(coefficients)
    
    # 6. Créer le fichier Excel avec moyennes
    print("\n📝 Calcul des moyennes en cours...")
    creer_fichier_excel_avec_moyennes(df, colonnes_notes, coefficients, fichier_sortie)
    
    # 7. Recharger pour afficher les statistiques
    df_final = pd.read_excel(fichier_sortie)
    afficher_statistiques(df_final, colonnes_notes)
    
    print("\n" + "="*70)
    print("✓ TRAITEMENT TERMINÉ AVEC SUCCÈS!")
    print("="*70)
    print(f"\n📁 Fichier créé: {fichier_sortie}")
    print("   - Moyenne calculée en Python (colonne jaune)")
    print("   - Moyenne avec formule Excel (colonne verte)")
    print("   - Coefficients affichés en ligne 2")
    print()

if __name__ == '__main__':
    programme_principal()
